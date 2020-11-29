


#################################-- Function Description --#################################

# Purpose:
#       This function reads an excel file for next functions
#    If a problem occurs warns the user and returns null

# INPUTS:
#   filePath      -> a string file name & path
#   sheetName     -> it would be better if the user give specific sheet name
#   willPrint     -> if the user wants to see the process or errors type True

# OUTPUTS
#   wb          -> Workbook
#   dataSheet   -> Data Sheet

#################################-- END Function Description --##############################

#################################-- IMPORT --#################################

# To control ".xlsx" extension
from GeneralLibrary import control_xlsx_extension

# import openpyxl to operate xlsx files
from openpyxl import load_workbook

#################################-- END  IMPORT --#############################



# this method reads values from given path
def read_excel_only_values(filePath,sheetName="veri",willPrint=False):

    # control if the name contains ".xlsx"
    filePath = control_xlsx_extension(filePath)

    # try to do process
    try:
        # get workbook - only values
        wb = load_workbook(filePath, data_only=True)

        # initialize dataheet
        dataSheet = None

        # iterate all sheet names
        for sheetName in wb.sheetnames:
            # control if it contains my name
            if sheetName in sheetName:

                # if the user wants to see result
                if (willPrint):
                    print("Last Selected Sheet: {}".format(sheetName))

                # get data sheet
                dataSheet = wb['veriler']

        # if the sheet exists
        if (dataSheet != None):
            # return them
            return wb,dataSheet
        # otherise
        else:
            # warn the user
            print("The sheet name : {} doesn't exist!!")
            # return null
            return None, None
    # In case any problem
    except:
        # warn the user
        print("There is an error on READ EXCEL ONLY VALUES Method !!")
        # return null
        return  None, None
















